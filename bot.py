import logging
import re
import os
import io
from datetime import datetime, date, timedelta

from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    ReplyKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardRemove,
)
from telegram.ext import (
    Application,
    MessageHandler,
    CommandHandler,
    CallbackQueryHandler,
    filters,
    ContextTypes,
)
from database import (
    init_db,
    add_transaction,
    update_transaction,
    delete_transaction,
    edit_transaction_comment,
    get_transaction_by_id,
    get_balance,
    get_recent_transactions,
    get_report,
    get_start_date,
    set_start_date,
    get_all_transactions,
    clear_all_transactions,
)

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
BOT_TOKEN     = os.environ["BOT_TOKEN"]
ADMIN_ID      = int(os.environ["ADMIN_ID"])
ALLOWED_GROUP = int(os.environ["ALLOWED_GROUP_ID"])

# Дополнительные пользователи через запятую: "111,222,333"
ALLOWED_USERS_RAW = os.environ.get("ALLOWED_USER_IDS", "")
ALLOWED_USERS = set()
if ALLOWED_USERS_RAW.strip():
    ALLOWED_USERS = {int(uid.strip()) for uid in ALLOWED_USERS_RAW.split(",") if uid.strip()}
ALLOWED_USERS.add(ADMIN_ID)


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def fmt(amount: int, currency: str) -> str:
    if currency == "UZS":
        return f"{abs(amount):,} UZS".replace(",", "\u00a0")
    return f"{abs(amount):,} $".replace(",", "\u00a0")


def parse_transaction(text: str):
    text = text.strip()
    if text.startswith("+"):
        t_type = "income"
        text = text[1:].strip()
    elif text.startswith("-"):
        t_type = "expense"
        text = text[1:].strip()
    else:
        t_type = "expense"

    match = re.search(r"\d[\d.]*", text)
    if not match:
        return None
    raw_number = match.group().replace(".", "")
    try:
        amount = int(raw_number)
    except ValueError:
        return None
    if amount <= 0:
        return None

    currency = "USD" if "$" in text else "UZS"
    rest = text[match.end():].replace("$", "").strip()
    return {"type": t_type, "amount": amount, "currency": currency, "comment": rest}


def parse_date_str(s: str):
    for fmt_str in ("%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(s.strip(), fmt_str).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def is_admin(uid: int) -> bool:
    return uid == ADMIN_ID


def is_allowed(uid: int) -> bool:
    return uid in ALLOWED_USERS


# ─────────────────────────────────────────────
# REPLY KEYBOARD (постоянные кнопки внизу экрана)
# ─────────────────────────────────────────────
def get_reply_keyboard(user_id: int) -> ReplyKeyboardMarkup:
    """Большие кнопки внизу экрана — возле клавиатуры."""
    if is_admin(user_id):
        buttons = [
            [KeyboardButton("💰 Баланс"),      KeyboardButton("📊 Отчёты")],
            [KeyboardButton("📥 Отчёт Excel"), KeyboardButton("🔧 Управление")],
            [KeyboardButton("➕ Добавить"),     KeyboardButton("📅 Дата начала")],
            [KeyboardButton("❓ Помощь")],
        ]
    else:
        buttons = [
            [KeyboardButton("💰 Баланс"),      KeyboardButton("📊 Отчёты")],
            [KeyboardButton("📥 Отчёт Excel"), KeyboardButton("❓ Помощь")],
        ]
    return ReplyKeyboardMarkup(buttons, resize_keyboard=True)


# ─────────────────────────────────────────────
# INLINE KEYBOARD для отчётов
# ─────────────────────────────────────────────
def reports_inline_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("📅 Сегодня",      callback_data="report:today"),
            InlineKeyboardButton("📆 Неделя",       callback_data="report:week"),
        ],
        [
            InlineKeyboardButton("🗓 Месяц",        callback_data="report:month"),
            InlineKeyboardButton("✏️ Период...",    callback_data="report:custom"),
        ],
        [
            InlineKeyboardButton("📊 Всё время",    callback_data="report:alltime"),
        ],
    ])


def admin_inline_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("➕ Добавить запись",    callback_data="admin:add"),
        ],
        [
            InlineKeyboardButton("🗑 Удалить запись",     callback_data="admin:delete"),
            InlineKeyboardButton("✏️ Изм. коммент",       callback_data="admin:edit"),
        ],
        [
            InlineKeyboardButton("🕐 Последние 10",       callback_data="admin:recent"),
        ],
        [
            InlineKeyboardButton("⚠️ Очистить все записи", callback_data="admin:clear_confirm"),
        ],
    ])


# ─────────────────────────────────────────────
# EXCEL ГЕНЕРАТОР
# ─────────────────────────────────────────────
def generate_excel(from_date: str, to_date: str, label: str) -> io.BytesIO:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl not installed")

    r    = get_report(from_date, to_date)
    txs  = r["transactions"]
    wb   = Workbook()

    # ── Лист 1: Транзакции ──
    ws = wb.active
    ws.title = "Транзакции"

    header_fill   = PatternFill("solid", start_color="1F4E79")
    income_fill   = PatternFill("solid", start_color="E2EFDA")
    expense_fill  = PatternFill("solid", start_color="FCE4D6")
    total_fill    = PatternFill("solid", start_color="D9E1F2")
    header_font   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    bold_font     = Font(bold=True, name="Arial", size=10)
    normal_font   = Font(name="Arial", size=10)
    center        = Alignment(horizontal="center", vertical="center")
    thin_border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    # Заголовок
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Финансовый отчёт: {label}"
    ws["A1"].font   = Font(bold=True, name="Arial", size=14, color="1F4E79")
    ws["A1"].alignment = center

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Период: {from_date} → {to_date}   |   Записей: {len(txs)}"
    ws["A2"].font      = Font(name="Arial", size=10, italic=True)
    ws["A2"].alignment = center

    # Шапка таблицы
    headers = ["#", "Дата", "Пользователь", "Сумма", "Валюта", "Тип", "Комментарий"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = thin_border

    # Данные
    for i, t in enumerate(txs, 1):
        row      = i + 4
        is_inc   = t["amount"] > 0
        row_fill = income_fill if is_inc else expense_fill
        sign     = "+" if is_inc else ""

        vals = [
            i,
            t["created_at"][:10],
            t["username"],
            f"{sign}{abs(t['amount']):,}".replace(",", " "),
            t["currency"],
            "Доход" if is_inc else "Расход",
            t["comment"] or "",
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font      = normal_font
            cell.fill      = row_fill
            cell.alignment = center if col != 7 else Alignment(vertical="center")
            cell.border    = thin_border

    # Итоговая строка
    total_row = len(txs) + 5
    ws.merge_cells(f"A{total_row}:C{total_row}")
    ws[f"A{total_row}"] = "ИТОГО:"
    ws[f"A{total_row}"].font      = bold_font
    ws[f"A{total_row}"].fill      = total_fill
    ws[f"A{total_row}"].alignment = center
    ws[f"A{total_row}"].border    = thin_border

    ws[f"D{total_row}"] = f"USD: +{r['income_usd']:,} / -{r['expense_usd']:,} = {r['balance_usd']:+,}".replace(",", " ")
    ws[f"D{total_row}"].font   = bold_font
    ws[f"D{total_row}"].fill   = total_fill
    ws[f"D{total_row}"].border = thin_border

    ws.merge_cells(f"E{total_row}:G{total_row}")
    ws[f"E{total_row}"] = f"UZS: +{r['income_uzs']:,} / -{r['expense_uzs']:,} = {r['balance_uzs']:+,}".replace(",", " ")
    ws[f"E{total_row}"].font      = bold_font
    ws[f"E{total_row}"].fill      = total_fill
    ws[f"E{total_row}"].alignment = center
    ws[f"E{total_row}"].border    = thin_border

    # Ширина колонок
    widths = [5, 12, 20, 18, 8, 10, 35]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[4].height = 20

    # ── Лист 2: Сводка ──
    ws2 = wb.create_sheet("Сводка")

    summary_data = [
        ("", "USD", "UZS"),
        ("📥 Доходы",  r["income_usd"],  r["income_uzs"]),
        ("📤 Расходы", r["expense_usd"], r["expense_uzs"]),
        ("💰 Баланс",  r["balance_usd"], r["balance_uzs"]),
    ]

    ws2["A1"] = f"Сводка: {label}"
    ws2["A1"].font = Font(bold=True, name="Arial", size=13, color="1F4E79")
    ws2.merge_cells("A1:C1")
    ws2["A1"].alignment = center

    for row_i, row_data in enumerate(summary_data, 3):
        for col_i, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_i, column=col_i, value=val)
            cell.border = thin_border
            cell.alignment = center
            if row_i == 3:
                cell.font = bold_font
                cell.fill = header_fill
                if col_i > 1:
                    cell.font = Font(bold=True, color="FFFFFF", name="Arial")
            elif row_i == 6:  # Баланс
                cell.fill = total_fill
                cell.font = bold_font
            else:
                cell.font = normal_font

    for col in ["A", "B", "C"]:
        ws2.column_dimensions[col].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────
# ТЕКСТЫ БАЛАНСА И ОТЧЁТА
# ─────────────────────────────────────────────
async def get_balance_text() -> str:
    bal    = get_balance()
    recent = get_recent_transactions(5)
    uzs    = bal.get("UZS", 0)
    usd    = bal.get("USD", 0)
    start  = get_start_date()

    start_line = f"\n📅 Учёт с: <b>{start}</b>" if start else ""
    lines = [
        f"💰 <b>Текущий баланс</b>{start_line}",
        "",
        f"💵 <b>USD:</b>  {'📈' if usd >= 0 else '📉'} {fmt(usd, 'USD')}",
        f"💳 <b>UZS:</b>  {'📈' if uzs >= 0 else '📉'} {fmt(uzs, 'UZS')}",
    ]
    if recent:
        lines += ["", "─────────────────", "🕐 <b>Последние 5 операций:</b>"]
        for r in recent:
            sign = "+" if r["amount"] > 0 else ""
            lines.append(
                f"  <b>#{r['id']}</b> {sign}{fmt(r['amount'], r['currency'])}"
                f" | {r['username']} | {r['comment'] or '—'}"
            )
    return "\n".join(lines)


def build_report_text(from_date, to_date, label) -> str:
    r = get_report(from_date, to_date)

    def si(val): return "📈" if val >= 0 else "📉"

    text = (
        f"📊 <b>Отчёт: {label}</b>\n"
        f"📅 {from_date} → {to_date} | Записей: {r['count']}\n\n"
        f"━━━━━━  💵 USD  ━━━━━━\n"
        f"📥 Доход:  +{fmt(r['income_usd'],  'USD')}\n"
        f"📤 Расход: -{fmt(r['expense_usd'], 'USD')}\n"
        f"{si(r['balance_usd'])} Итого:  {'+' if r['balance_usd'] >= 0 else ''}{fmt(r['balance_usd'], 'USD')}\n\n"
        f"━━━━━━  💳 UZS  ━━━━━━\n"
        f"📥 Доход:  +{fmt(r['income_uzs'],  'UZS')}\n"
        f"📤 Расход: -{fmt(r['expense_uzs'], 'UZS')}\n"
        f"{si(r['balance_uzs'])} Итого:  {'+' if r['balance_uzs'] >= 0 else ''}{fmt(r['balance_uzs'], 'UZS')}\n"
    )
    txs = r["transactions"][-10:]
    if txs:
        text += "\n─────────────────\n<b>Записи периода:</b>\n"
        for t in reversed(txs):
            sign = "+" if t["amount"] > 0 else ""
            dt   = t["created_at"][5:10]
            text += f"  <b>#{t['id']}</b> {dt} {sign}{fmt(t['amount'], t['currency'])} | {t['comment'] or '—'}\n"
    return text


# ─────────────────────────────────────────────
# ГРУППА: новые сообщения
# ─────────────────────────────────────────────
async def handle_group_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message or not update.message.text:
        return
    msg       = update.message
    user      = msg.from_user
    full_text = msg.text.strip()

    if msg.chat.id != ALLOWED_GROUP:
        return

    user_display = f"@{user.username}" if user.username else user.full_name

    # Каждая строка — отдельная транзакция
    lines  = [l.strip() for l in full_text.splitlines() if l.strip()]
    saved  = []   # (tx_id, tx)
    errors = []   # некорректные строки (без цифр)

    for line in lines:
        tx = parse_transaction(line)
        if tx is None:
            if not re.search(r"\d", line):
                errors.append(line)
            continue

        sign  = 1 if tx["type"] == "income" else -1
        tx_id = add_transaction(
            user_id=user.id,
            username=user_display,
            amount=sign * tx["amount"],
            currency=tx["currency"],
            comment=tx["comment"],
            raw_text=line,
            msg_id=msg.message_id,
        )
        if tx_id != -1:
            saved.append((tx_id, tx))

    if not saved and not errors:
        return

    # Одно уведомление на всё сообщение
    out = []
    if saved:
        n = len(saved)
        word = "запись" if n == 1 else ("записи" if n < 5 else "записей")
        out.append(f"📋 <b>{n} {word} от</b> {user_display}:\n")
        delete_buttons = []
        for tx_id, tx in saved:
            sign_str = "+" if tx["type"] == "income" else "-"
            icon     = "📥" if tx["type"] == "income" else "📤"
            out.append(
                f"{icon} <b>#{tx_id}</b>  {sign_str}{fmt(tx['amount'], tx['currency'])}"
                f"  |  {tx['comment'] or '—'}"
            )
            delete_buttons.append(InlineKeyboardButton(f"🗑 #{tx_id}", callback_data=f"del:{tx_id}"))
        keyboard = InlineKeyboardMarkup([delete_buttons[i:i+3] for i in range(0, len(delete_buttons), 3)])
    else:
        keyboard = None

    if errors:
        out.append("\n⚠️ Некорректные строки:")
        for e in errors:
            out.append(f"  <code>{e}</code>")

    await context.bot.send_message(
        chat_id=ADMIN_ID,
        text="\n".join(out),
        parse_mode="HTML",
        reply_markup=keyboard,
    )


# ─────────────────────────────────────────────
# ГРУППА: редактирование сообщений
# ─────────────────────────────────────────────
async def handle_group_edit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Пользователь отредактировал сообщение в группе."""
    if not update.edited_message or not update.edited_message.text:
        return
    msg  = update.edited_message
    user = msg.from_user
    text = msg.text.strip()

    if msg.chat.id != ALLOWED_GROUP:
        return

    user_display = f"@{user.username}" if user.username else user.full_name

    # Ищем запись по message_id
    from database import get_transaction_by_msg_id
    old_tx = get_transaction_by_msg_id(msg.message_id)

    tx = parse_transaction(text)

    if old_tx is None:
        # Записи не было — пробуем создать новую
        if tx:
            sign  = 1 if tx["type"] == "income" else -1
            tx_id = add_transaction(
                user_id=user.id, username=user_display,
                amount=sign * tx["amount"], currency=tx["currency"],
                comment=tx["comment"], raw_text=text, msg_id=msg.message_id,
            )
            await context.bot.send_message(
                chat_id=ADMIN_ID,
                text=(
                    f"✏️ <b>Редактирование → новая запись #{tx_id}</b>\n"
                    f"👤 {user_display}\n"
                    f"💰 {'+'if tx['type']=='income' else '-'}{fmt(tx['amount'], tx['currency'])}\n"
                    f"📝 {tx['comment'] or '—'}"
                ),
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton(f"🗑 Удалить #{tx_id}", callback_data=f"del:{tx_id}")]
                ]),
            )
        return

    # Запись существует
    if tx is None:
        # Новый текст нечитаемый — уведомляем
        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=(
                f"⚠️ <b>Редактирование: некорректный текст</b>\n"
                f"👤 {user_display} изменил запись <b>#{old_tx['id']}</b>\n"
                f"Старое: <code>{old_tx['raw_text']}</code>\n"
                f"Новое:  <code>{text}</code>\n\n"
                f"Запись <b>не изменена</b> в БД."
            ),
            parse_mode="HTML",
        )
        return

    # Обновляем запись в БД
    sign      = 1 if tx["type"] == "income" else -1
    old_sign  = "+" if old_tx["amount"] > 0 else ""
    new_sign  = "+" if tx["type"] == "income" else "-"

    update_transaction(
        tx_id=old_tx["id"],
        amount=sign * tx["amount"],
        currency=tx["currency"],
        comment=tx["comment"],
        raw_text=text,
    )

    await context.bot.send_message(
        chat_id=ADMIN_ID,
        text=(
            f"✏️ <b>Запись #{old_tx['id']} обновлена</b>\n"
            f"👤 {user_display}\n"
            f"Было:  {old_sign}{fmt(old_tx['amount'], old_tx['currency'])} | {old_tx['comment'] or '—'}\n"
            f"Стало: {new_sign}{fmt(tx['amount'], tx['currency'])} | {tx['comment'] or '—'}"
        ),
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton(f"🗑 Удалить #{old_tx['id']}", callback_data=f"del:{old_tx['id']}")]
        ]),
    )


# ─────────────────────────────────────────────
# /start — показываем Reply-клавиатуру
# ─────────────────────────────────────────────
async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    if not is_allowed(user.id):
        await update.message.reply_text("⛔ У вас нет доступа.")
        return

    greeting = "👋 Привет, <b>Администратор</b>!" if is_admin(user.id) else f"👋 Привет, <b>{user.first_name}</b>!"
    await update.message.reply_text(
        f"{greeting}\n\nКнопки меню появились внизу экрана 👇",
        parse_mode="HTML",
        reply_markup=get_reply_keyboard(user.id),
    )


# ─────────────────────────────────────────────
# ОБРАБОТКА REPLY-КНОПОК
# ─────────────────────────────────────────────
async def handle_reply_buttons(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    if not is_allowed(user.id):
        return

    text = update.message.text.strip()

    # ── 💰 Баланс ──
    if text == "💰 Баланс":
        balance_text = await get_balance_text()
        await update.message.reply_text(balance_text, parse_mode="HTML")
        return

    # ── 📊 Отчёты ──
    if text == "📊 Отчёты":
        await update.message.reply_text(
            "📊 <b>Выберите период:</b>",
            parse_mode="HTML",
            reply_markup=reports_inline_keyboard(),
        )
        return

    # ── 📥 Отчёт Excel ──
    if text == "📥 Отчёт Excel":
        await update.message.reply_text(
            "📥 <b>Отчёт Excel — выберите период:</b>",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("📅 Сегодня",   callback_data="excel:today"),
                    InlineKeyboardButton("📆 Неделя",    callback_data="excel:week"),
                ],
                [
                    InlineKeyboardButton("🗓 Месяц",     callback_data="excel:month"),
                    InlineKeyboardButton("✏️ Период...", callback_data="excel:custom"),
                ],
                [
                    InlineKeyboardButton("📊 Всё время", callback_data="excel:alltime"),
                ],
            ]),
        )
        return

    # ── 🔧 Управление (только админ) ──
    if text == "🔧 Управление":
        if not is_admin(user.id):
            return
        recent = get_recent_transactions(3)
        lines  = ["🔧 <b>Панель администратора</b>\n"]
        if recent:
            lines.append("🕐 <b>Последние записи:</b>")
            for r in recent:
                sign = "+" if r["amount"] > 0 else ""
                lines.append(f"  <b>#{r['id']}</b> {sign}{fmt(r['amount'], r['currency'])} | {r['comment'] or '—'}")
        await update.message.reply_text(
            "\n".join(lines),
            parse_mode="HTML",
            reply_markup=admin_inline_keyboard(),
        )
        return

    # ── ➕ Добавить запись (только админ) ──
    if text == "➕ Добавить":
        if not is_admin(user.id):
            return
        context.user_data["awaiting"] = "manual_add"
        await update.message.reply_text(
            "➕ <b>Добавление записи вручную</b>\n\n"
            "Введите запись в том же формате что и в группе:\n\n"
            "  <code>+600$ коммент</code>  → доход USD\n"
            "  <code>+500000 коммент</code> → доход UZS\n"
            "  <code>-150000 аренда</code>  → расход UZS\n\n"
            "Можно несколько строк сразу:\n"
            "  <code>-56000 gaz\n-32000 otopark\n-160000 doktor</code>\n\n"
            "Или нажмите /start для отмены.",
            parse_mode="HTML",
        )
        return

    # ── 📅 Дата начала (только админ) ──
    if text == "📅 Дата начала":
        if not is_admin(user.id):
            return
        current = get_start_date()
        current_str = f"<b>{current}</b>" if current else "<i>не задана</i>"
        context.user_data["awaiting"] = "setstart"
        await update.message.reply_text(
            f"📅 <b>Дата начала учёта</b>\n\n"
            f"Текущая: {current_str}\n\n"
            f"Введите новую дату: <code>01.07.2025</code>\n"
            f"Или напишите <code>сброс</code> чтобы учитывать все записи.",
            parse_mode="HTML",
        )
        return

    # ── ❓ Помощь ──
    if text == "❓ Помощь":
        help_text = (
            "📋 <b>Справка</b>\n\n"
            "<b>Записи в группе:</b>\n"
            "  <code>+600$ коммент</code>  → доход USD\n"
            "  <code>+500000 коммент</code> → доход UZS\n"
            "  <code>-150000 аренда</code>  → расход UZS\n"
            "  <code>50000</code>           → расход UZS\n\n"
            "<b>Редактирование:</b> если изменить сообщение в группе — бот автоматически обновит запись и уведомит администратора.\n"
        )
        if is_admin(user.id):
            help_text += (
                "\n<b>Управление записями:</b>\n"
                "Кнопка 🔧 Управление → удалить или изменить комментарий\n"
                "Кнопка под уведомлением → быстрое удаление\n\n"
                "<b>Добавить пользователя:</b>\n"
                "Railway → Variables → ALLOWED_USER_IDS\n"
                "Добавить ID через запятую: <code>123456,789012</code>\n"
            )
        await update.message.reply_text(help_text, parse_mode="HTML")
        return

    # ── Режим ожидания ввода ──
    await handle_awaiting_input(update, context)


async def handle_awaiting_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user     = update.effective_user
    awaiting = context.user_data.get("awaiting")
    text     = update.message.text.strip()

    if not awaiting:
        return

    today = date.today()

    # ── Ручное добавление записи (только админ) ──
    if awaiting == "manual_add":
        context.user_data.pop("awaiting", None)
        if not is_admin(user.id):
            return

        lines  = [l.strip() for l in text.splitlines() if l.strip()]
        saved  = []
        failed = []

        for line in lines:
            tx = parse_transaction(line)
            if tx is None:
                failed.append(line)
                continue
            sign  = 1 if tx["type"] == "income" else -1
            tx_id = add_transaction(
                user_id=user.id,
                username=f"@{user.username}" if user.username else user.full_name,
                amount=sign * tx["amount"],
                currency=tx["currency"],
                comment=tx["comment"],
                raw_text=line,
                msg_id=None,
            )
            if tx_id != -1:
                saved.append((tx_id, tx))
            else:
                failed.append(f"{line} (до даты начала учёта)")

        out = []
        if saved:
            out.append(f"✅ <b>Добавлено {len(saved)} запис{'ь' if len(saved)==1 else 'и' if len(saved)<5 else 'ей'}:</b>\n")
            del_btns = []
            for tx_id, tx in saved:
                s = "+" if tx["type"] == "income" else "-"
                icon = "📥" if tx["type"] == "income" else "📤"
                out.append(f"{icon} <b>#{tx_id}</b>  {s}{fmt(tx['amount'], tx['currency'])}  |  {tx['comment'] or '—'}")
                del_btns.append(InlineKeyboardButton(f"🗑 #{tx_id}", callback_data=f"del:{tx_id}"))
            keyboard = InlineKeyboardMarkup([del_btns[i:i+3] for i in range(0, len(del_btns), 3)])
        else:
            keyboard = None

        if failed:
            out.append(f"\n❌ <b>Не распознано:</b>")
            for f in failed:
                out.append(f"  <code>{f}</code>")

        await update.message.reply_text("\n".join(out), parse_mode="HTML", reply_markup=keyboard)
        return

    # ── Кастомный период (отчёт или Excel) ──
    if awaiting in ("custom_report", "custom_excel"):
        context.user_data.pop("awaiting", None)
        if "-" not in text:
            await update.message.reply_text("❌ Формат: <code>01.06-30.06</code>", parse_mode="HTML")
            return
        parts = text.split("-")
        def norm(d):
            return d + f".{today.year}" if d.count(".") == 1 else d
        fd = parse_date_str(norm(parts[0]))
        td = parse_date_str(norm(parts[1]))
        if not fd or not td:
            await update.message.reply_text("❌ Неверная дата. Пример: <code>01.06-30.06</code>", parse_mode="HTML")
            return
        label = f"{parts[0].strip()} — {parts[1].strip()}"
        if awaiting == "custom_excel":
            await send_excel(update, fd, td, label)
        else:
            await update.message.reply_text(build_report_text(fd, td, label), parse_mode="HTML")
        return

    # ── Удаление: ввод ID ──
    if awaiting == "delete_id":
        context.user_data.pop("awaiting", None)
        if not is_admin(user.id): return
        try:
            tx_id = int(text)
        except ValueError:
            await update.message.reply_text("❌ ID должен быть числом.")
            return
        tx = get_transaction_by_id(tx_id)
        if not tx:
            await update.message.reply_text(f"❌ Запись #{tx_id} не найдена.")
            return
        sign = "+" if tx["amount"] > 0 else ""
        await update.message.reply_text(
            f"🗑 <b>Удалить эту запись?</b>\n\n"
            f"<b>#{tx['id']}</b> | {sign}{fmt(tx['amount'], tx['currency'])}\n"
            f"👤 {tx['username']} | 📝 {tx['comment'] or '—'}",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("✅ Да, удалить", callback_data=f"del:{tx_id}"),
                    InlineKeyboardButton("❌ Отмена",      callback_data="noop"),
                ]
            ]),
        )
        return

    # ── Редактирование: ввод ID ──
    if awaiting == "edit_id":
        if not is_admin(user.id): return
        try:
            tx_id = int(text)
        except ValueError:
            context.user_data.pop("awaiting", None)
            await update.message.reply_text("❌ ID должен быть числом.")
            return
        tx = get_transaction_by_id(tx_id)
        if not tx:
            context.user_data.pop("awaiting", None)
            await update.message.reply_text(f"❌ Запись #{tx_id} не найдена.")
            return
        context.user_data["awaiting"]   = "edit_text"
        context.user_data["edit_tx_id"] = tx_id
        sign = "+" if tx["amount"] > 0 else ""
        await update.message.reply_text(
            f"✏️ <b>Запись #{tx_id}</b>: {sign}{fmt(tx['amount'], tx['currency'])}\n"
            f"Текущий коммент: <i>{tx['comment'] or '—'}</i>\n\n"
            f"Введите <b>новый комментарий</b>:",
            parse_mode="HTML",
        )
        return

    # ── Редактирование: ввод нового комментария ──
    if awaiting == "edit_text":
        if not is_admin(user.id): return
        context.user_data.pop("awaiting", None)
        tx_id = context.user_data.pop("edit_tx_id", None)
        if not tx_id: return
        edit_transaction_comment(tx_id, text)
        await update.message.reply_text(
            f"✅ <b>Запись #{tx_id} обновлена</b>\nКомментарий: <i>{text}</i>",
            parse_mode="HTML",
        )
        return

    # ── Дата начала учёта ──
    if awaiting == "setstart":
        if not is_admin(user.id): return
        context.user_data.pop("awaiting", None)
        if text.lower() in ("сброс", "reset", "off"):
            set_start_date("")
            await update.message.reply_text("✅ Дата начала сброшена. Учитываются все записи.")
            return
        parsed = parse_date_str(text)
        if not parsed:
            await update.message.reply_text("❌ Неверный формат. Пример: <code>01.07.2025</code>", parse_mode="HTML")
            return
        set_start_date(parsed)
        await update.message.reply_text(
            f"✅ <b>Дата начала учёта: {parsed}</b>\nЗаписи до этой даты игнорируются.",
            parse_mode="HTML",
        )
        return


# ─────────────────────────────────────────────
# ОТПРАВКА EXCEL
# ─────────────────────────────────────────────
async def send_excel(update_or_query, from_date, to_date, label):
    """Генерирует и отправляет Excel файл."""
    # Определяем метод отправки
    if hasattr(update_or_query, 'message') and update_or_query.message:
        send = update_or_query.message.reply_document
        chat_id = None
    else:
        # CallbackQuery
        send = None
        chat_id = update_or_query.message.chat_id
        bot    = update_or_query.get_bot() if hasattr(update_or_query, 'get_bot') else None

    try:
        buf      = generate_excel(from_date, to_date, label)
        filename = f"report_{from_date}_{to_date}.xlsx"
        caption  = f"📊 Отчёт: {label}\n📅 {from_date} → {to_date}"

        if send:
            await send(document=buf, filename=filename, caption=caption)
        else:
            await update_or_query.bot.send_document(
                chat_id=update_or_query.message.chat_id,
                document=buf,
                filename=filename,
                caption=caption,
            )
    except Exception as e:
        logger.error(f"Excel generation error: {e}")
        msg = "❌ Ошибка генерации Excel. Убедитесь что openpyxl установлен."
        if send:
            await send(document=None) if False else None
            # fallback
            await update_or_query.message.reply_text(msg)
        else:
            await update_or_query.bot.send_message(chat_id=update_or_query.message.chat_id, text=msg)


# ─────────────────────────────────────────────
# CALLBACK HANDLER
# ─────────────────────────────────────────────
async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query   = update.callback_query
    user_id = query.from_user.id
    data    = query.data
    await query.answer()

    today = date.today()

    # ── Noop (отмена) ──
    if data == "noop":
        await query.edit_message_text("❌ Отменено.")
        return

    # ── Быстрое удаление ──
    if data.startswith("del:"):
        if not is_admin(user_id):
            await query.answer("⛔ Только администратор.", show_alert=True)
            return
        tx_id = int(data.split(":")[1])
        tx    = get_transaction_by_id(tx_id)
        if not tx:
            await query.edit_message_text("❌ Запись не найдена (уже удалена?).")
            return
        delete_transaction(tx_id)
        sign = "+" if tx["amount"] > 0 else ""
        await query.edit_message_text(
            f"🗑 <b>Запись #{tx_id} удалена</b>\n"
            f"{sign}{fmt(tx['amount'], tx['currency'])} | {tx['comment'] or '—'}",
            parse_mode="HTML",
        )
        return

    # ── Отчёты inline ──
    if data.startswith("report:"):
        if not is_allowed(user_id): return
        period = data.split(":")[1]
        if period == "today":
            fd = td = today.strftime("%Y-%m-%d")
            label = f"Сегодня ({today.strftime('%d.%m.%Y')})"
        elif period == "week":
            fd    = (today - timedelta(days=today.weekday())).strftime("%Y-%m-%d")
            td    = today.strftime("%Y-%m-%d")
            label = "Текущая неделя"
        elif period == "month":
            fd    = today.strftime("%Y-%m-01")
            td    = today.strftime("%Y-%m-%d")
            label = f"Текущий месяц ({today.strftime('%m.%Y')})"
        elif period == "alltime":
            # От даты начала учёта (или самой первой записи) до сегодня
            start = get_start_date()
            if start:
                fd    = start
                label = f"Всё время (с {start})"
            else:
                # Берём дату самой первой записи в БД
                from database import get_first_transaction_date
                first = get_first_transaction_date()
                fd    = first if first else today.strftime("%Y-%m-%d")
                label = f"Всё время (с {fd})"
            td = today.strftime("%Y-%m-%d")
        elif period == "custom":
            context.user_data["awaiting"] = "custom_report"
            await query.edit_message_text(
                "✏️ Введите период:\n\n<code>01.06-30.06</code>\nили\n<code>01.06.2025-30.06.2025</code>",
                parse_mode="HTML",
            )
            return
        else:
            return
        await query.edit_message_text(build_report_text(fd, td, label), parse_mode="HTML")
        return

    # ── Excel inline ──
    if data.startswith("excel:"):
        if not is_allowed(user_id): return
        period = data.split(":")[1]
        if period == "today":
            fd = td = today.strftime("%Y-%m-%d")
            label = f"Сегодня ({today.strftime('%d.%m.%Y')})"
        elif period == "week":
            fd    = (today - timedelta(days=today.weekday())).strftime("%Y-%m-%d")
            td    = today.strftime("%Y-%m-%d")
            label = "Неделя"
        elif period == "month":
            fd    = today.strftime("%Y-%m-01")
            td    = today.strftime("%Y-%m-%d")
            label = f"Месяц {today.strftime('%m.%Y')}"
        elif period == "alltime":
            start = get_start_date()
            if start:
                fd    = start
                label = f"Всё время (с {start})"
            else:
                from database import get_first_transaction_date
                first = get_first_transaction_date()
                fd    = first if first else today.strftime("%Y-%m-%d")
                label = f"Всё время (с {fd})"
            td = today.strftime("%Y-%m-%d")
        elif period == "custom":
            context.user_data["awaiting"] = "custom_excel"
            await query.edit_message_text(
                "✏️ Введите период:\n\n<code>01.06-30.06</code>",
                parse_mode="HTML",
            )
            return
        else:
            return
        await query.edit_message_text(f"⏳ Генерирую Excel...")
        try:
            buf      = generate_excel(fd, td, label)
            filename = f"report_{fd}_{td}.xlsx"
            await query.message.reply_document(
                document=buf,
                filename=filename,
                caption=f"📊 {label} | {fd} → {td}",
            )
        except Exception as e:
            logger.error(f"Excel error: {e}")
            await query.message.reply_text("❌ Ошибка. Убедитесь что openpyxl установлен.")
        return

    # ── Управление ──
    if data == "admin:recent":
        if not is_admin(user_id): return
        recent = get_recent_transactions(10)
        lines  = ["🕐 <b>Последние 10 записей:</b>\n"]
        for r in recent:
            sign = "+" if r["amount"] > 0 else ""
            dt   = r["created_at"][5:10]
            lines.append(f"<b>#{r['id']}</b> {dt}  {sign}{fmt(r['amount'], r['currency'])}  | {r['username']}  | {r['comment'] or '—'}")
        await query.edit_message_text("\n".join(lines), parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Назад", callback_data="admin:back")]]))
        return

    if data == "admin:add":
        if not is_admin(user_id): return
        context.user_data["awaiting"] = "manual_add"
        await query.edit_message_text(
            "➕ <b>Добавление записи вручную</b>\n\n"
            "Введите в том же формате что и в группе.\n"
            "Можно несколько строк сразу:\n\n"
            "<code>-56000 gaz\n-32000 otopark\n+1000000 оплата</code>\n\n"
            "Или нажмите /start для отмены.",
            parse_mode="HTML",
        )
        return

    if data == "admin:back":
        if not is_admin(user_id): return
        await query.edit_message_text("🔧 <b>Панель администратора</b>", parse_mode="HTML",
            reply_markup=admin_inline_keyboard())
        return

    if data == "admin:clear_confirm":
        if not is_admin(user_id): return
        from database import get_first_transaction_date
        first = get_first_transaction_date()
        bal   = get_balance()
        uzs   = bal.get("UZS", 0)
        usd   = bal.get("USD", 0)
        await query.edit_message_text(
            f"⚠️ <b>Удалить ВСЕ записи?</b>\n\n"
            f"Это действие нельзя отменить!\n\n"
            f"Текущий баланс будет потерян:\n"
            f"💵 {fmt(usd, 'USD')}\n"
            f"💳 {fmt(uzs, 'UZS')}\n"
            f"📅 Записи с: {first or '—'}",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("✅ Да, удалить всё", callback_data="admin:clear_execute"),
                    InlineKeyboardButton("❌ Отмена",           callback_data="admin:back"),
                ]
            ]),
        )
        return

    if data == "admin:clear_execute":
        if not is_admin(user_id): return
        await query.edit_message_text("⏳ Удаляю все записи...")
        count = clear_all_transactions()
        if count >= 0:
            await query.edit_message_text(
                f"✅ <b>Удалено {count} записей.</b>\n\nБаза данных очищена.",
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("◀️ Назад", callback_data="admin:back")]
                ]),
            )
        else:
            await query.edit_message_text(
                "❌ Ошибка при очистке. Попробуйте ещё раз.",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("◀️ Назад", callback_data="admin:back")]
                ]),
            )
        return

    if data == "admin:delete":
        if not is_admin(user_id): return
        context.user_data["awaiting"] = "delete_id"
        await query.edit_message_text(
            "🗑 Введите <b>ID записи</b> для удаления:\n(ID виден в уведомлениях и списке записей)",
            parse_mode="HTML",
        )
        return

    if data == "admin:edit":
        if not is_admin(user_id): return
        context.user_data["awaiting"] = "edit_id"
        await query.edit_message_text(
            "✏️ Введите <b>ID записи</b> для изменения комментария:",
            parse_mode="HTML",
        )
        return


# ─────────────────────────────────────────────
# ЗАПУСК
# ─────────────────────────────────────────────
def main():
    init_db()
    app = Application.builder().token(BOT_TOKEN).build()

    private = filters.ChatType.PRIVATE
    groups  = filters.ChatType.GROUPS

    app.add_handler(CommandHandler("start", start_command, filters=private))

    # Reply-кнопки и ввод текста в ЛС
    app.add_handler(MessageHandler(private & filters.TEXT, handle_reply_buttons))

    # Новые сообщения в группе
    app.add_handler(MessageHandler(groups & filters.TEXT & ~filters.UpdateType.EDITED_MESSAGE, handle_group_message))

    # Редактирование сообщений в группе
    app.add_handler(MessageHandler(groups & filters.UpdateType.EDITED_MESSAGE, handle_group_edit))

    app.add_handler(CallbackQueryHandler(handle_callback))

    logger.info("Bot v4 started.")
    app.run_polling(drop_pending_updates=True, allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
